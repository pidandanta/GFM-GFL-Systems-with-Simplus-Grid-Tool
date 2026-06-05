import openpyxl
import os
import shutil

# 【路径修复】：自动获取当前脚本所在的绝对路径
project_dir = os.path.dirname(os.path.abspath(__file__))
base_file = os.path.join(project_dir, 'NETS_NYPS_68Bus.xlsx')

# 定义 3 种方案的【额外 3 台 GFM】的安装位置
# 基础 GFM 已经占据了 1, 2, 3, 4
schemes = {
    'Scheme1_Uniform': [6, 10, 12],  # 均匀打散
    'Scheme2_Hsys': [10, 11, 12],  # 靠近 13-16 同步机骨干
    'Scheme3_gCSR': [7, 8, 9]  # 你的算法算出的三大命门
}

# 变流器参数 (沿用你之前的大容量自适应+强阻尼维稳参数)
gfm_params = [0.2, 0.05, 0.002, 0.001, 0.5, 0.001, 0.0005, 0, 20, 100]
gfl_params = [1.0, 0.5, 0.002, 0.001, 0.05, 0.1, 250]

print("🚀 开始生成 3 种对比方案的 Excel 拓扑文件...")

for scheme_name, extra_gfm in schemes.items():
    new_filename = f'NETS_68Bus_{scheme_name}.xlsx'
    target_path = os.path.join(project_dir, new_filename)
    shutil.copy(base_file, target_path)

    wb = openpyxl.load_workbook(target_path)

    # 1. 关闭导纳频谱图绘制防崩溃
    ws_basic = wb['Basic']
    for row in ws_basic.iter_rows(min_row=1, max_col=2):
        if row[0].value and 'admittance' in str(row[0].value).lower():
            row[1].value = 0

    # 2. 定向投放 GFM 资源
    ws_app = wb['Apparatus']

    # 结合基础 4 台，计算出该方案下所有的 GFM 节点编号
    target_gfm_buses = [1, 2, 3, 4] + extra_gfm

    for row in ws_app.iter_rows(min_row=3, max_col=18):
        bus_no = row[0].value
        if bus_no is None: continue

        # 保护 13-16 同步机骨干
        if 13 <= bus_no <= 16:
            row[1].value = 0

        # 对 1-12 内部节点进行 GFM/GFL 区分
        elif 1 <= bus_no <= 12:
            if bus_no in target_gfm_buses:
                row[1].value = 20  # GFM
                for i in range(2, 18): row[i].value = None
                for i, val in enumerate(gfm_params): row[2 + i].value = val
            else:
                row[1].value = 11  # GFL
                for i in range(2, 18): row[i].value = None
                for i, val in enumerate(gfl_params): row[2 + i].value = val

    wb.save(target_path)
    print(f"✅ 成功生成 {new_filename} (GFM 安装在: {target_gfm_buses})")

print("🎉 3 种拓扑文件已全部就绪！")